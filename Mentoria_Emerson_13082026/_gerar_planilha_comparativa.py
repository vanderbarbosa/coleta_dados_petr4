# -*- coding: utf-8 -*-
# ==============================================================================
#   Gera a planilha comparativa das pesquisas levantadas para a mentoria
#   Saída: Mentoria_Emerson_13082026/PLANILHA_COMPARATIVA_PESQUISAS.xlsx
#
#   Três abas:
#     1. Comparativo   — uma linha por pesquisa, 18 colunas
#     2. Nossos números — o que a nossa pesquisa mede hoje, com as correções
#     3. Como usar     — o plano de adaptação, por prioridade
# ==============================================================================
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

AQUI = Path(__file__).resolve().parent
SAIDA = AQUI / "PLANILHA_COMPARATIVA_PESQUISAS.xlsx"

# paleta sóbria, legível em impressão
AZUL = "1F3864"
CINZA_CAB = "D9E2F3"
VERDE = "E2EFDA"      # relevância alta
AMARELO = "FFF2CC"    # média
CINZA = "F2F2F2"      # baixa
VERMELHO = "FCE4E4"   # atenção / ressalva forte
NOSSO = "FFE699"      # a nossa pesquisa

BORDA = Border(*[Side(style="thin", color="BFBFBF")] * 4)

COLUNAS = [
    ("#", 5),
    ("Autor(es)", 26),
    ("Ano", 6),
    ("Título (abreviado)", 42),
    ("Veículo", 26),
    ("Idioma do corpus", 14),
    ("Mercado / Ativo", 24),
    ("Volume de textos", 18),
    ("Período", 14),
    ("Encoder / Tecnologia", 34),
    ("Alvo da previsão", 26),
    ("Modelo preditivo / referência", 30),
    ("Resultado principal", 52),
    ("Teste de significância", 24),
    ("Código público", 14),
    ("Relevância", 12),
    ("Como usar na nossa pesquisa", 58),
    ("Ressalvas", 52),
]

# ── as pesquisas ──────────────────────────────────────────────────────────────
# relevancia: ALTA / MEDIA / BAIXA / NOSSA / ATENCAO
LINHAS = [
    # ---------------------------------------------------------------- a nossa
    dict(rel="NOSSA",
         autor="Silva, V. B.", ano="2026",
         titulo="O impacto do sentimento de notícias financeiras na previsão de direção e volatilidade da PETR4",
         veiculo="Dissertação de mestrado, PUCPR/PPGIa",
         idioma="Português", mercado="B3 — PETR4 (ação individual)",
         volume="205.697 manchetes", periodo="2018–2025",
         tec="FinBERT-PT-BR (BERTimbau adaptado)",
         alvo="direção do preço e nível da volatilidade",
         modelo="XGBoost e SVM; GARCH(1,1)-t; HAR; regressão quantílica",
         resultado="Direção 54,5% (ganho de 4,4 p.p. sobre preços). Volatilidade: coeficiente −0,2924 (p=0,0002), mas NÃO supera o HAR. Efeito é de cauda.",
         teste="binomial, McNemar, Diebold-Mariano, bootstrap pareado",
         codigo="repositório próprio",
         usar="—",
         ressalva="NÃO usar o número 54,93% da ponderação por confiança: é acurácia de validação; no teste dá 50,31% contra 53,88% sem ponderar (Seção 4.n)."),

    # ------------------------------------------- fazem o mesmo que nós (ALTA)
    dict(rel="ALTA",
         autor="Hashamia, R.; Maldonado, F.", ano="2025",
         titulo="Can News Predict the Direction of Oil Price Volatility? A Language Model Approach with SHAP",
         veiculo="arXiv:2508.20707",
         idioma="Inglês", mercado="Futuros de Brent (petróleo)",
         volume="592.858 manchetes (Reuters)", periodo="2014–2023",
         tec="VADER, TextBlob, FinBERT, CrudeBERT; embeddings GloVe, FastText, BERT, Gemini, LLaMA",
         alvo="DIREÇÃO da volatilidade do dia seguinte (binária)",
         modelo="Regressão logística, Naive Bayes, kNN; referência HAR",
         resultado="A contagem de notícias superou as medidas de sentimento. FastText foi o melhor embedding, acima das cabeças de sentimento. Estratificam 4 regimes de mercado.",
         teste="McNemar",
         codigo="SIM — github.com/Romina-Hashami/Textual_Direction_Prediction_Oil_Volatility",
         usar="PRIORIDADE 1. Adotar o alvo DIREÇÃO DA VOLATILIDADE — via do meio entre direção do preço (≈acaso) e nível da volatilidade (HAR imbatível). Nunca testamos. Também: testar contagem de notícias contra esse alvo.",
         ressalva="Não foi possível ler as tabelas de desempenho no PDF; os achados vêm do README do repositório."),

    dict(rel="ALTA",
         autor="Bodilsen, S. T.; Lunde, A.", ano="2025",
         titulo="Exploiting News Analytics for Volatility Forecasting",
         veiculo="Journal of Applied Econometrics, 40(1):18–36",
         idioma="Inglês", mercado="Ações individuais dos EUA e S&P 500",
         volume="não recuperado", periodo="não recuperado",
         tec="analítica de notícias comercial (news analytics)",
         alvo="volatilidade realizada, vários horizontes",
         modelo="família HAR",
         resultado="Notícia ESPECÍFICA DA EMPRESA não acrescenta ao HAR. Notícia MACROECONÔMICA melhora de forma significativa, sobretudo em HORIZONTES LONGOS.",
         teste="comparação preditiva fora da amostra",
         codigo="não",
         usar="JÁ TESTADO na PETR4 (Seção 4.o): a hipótese INVERTE-SE. Aqui a notícia da empresa ajuda (+1,77% em 22 dias, p=0,0574) e a macro PIORA significativamente (p=0,015 e p=0,020). A metade sobre o horizonte confirma-se.",
         ressalva="Periódico de primeira linha; Lunde é coautor de Hansen e Lunde (2005). Texto integral bloqueado (HTTP 403): dados via resumo."),

    dict(rel="ALTA",
         autor="Halousková, M.; Lyócsa, Š.", ano="2025",
         titulo="Forecasting U.S. equity market volatility with attention and sentiment to the economy",
         veiculo="arXiv:2503.19767",
         idioma="Inglês", mercado="404 ações do S&P 500",
         volume="WSJ, FT, Google Trends, Wikipédia, Twitter", periodo="2010–2021",
         tec="FinBERT",
         alvo="variância realizada (5 minutos)",
         modelo="HAR, HAR-M, CSR-HAR; subconjuntos completos e LASSO adaptativo",
         resultado="SUPERAM o HAR em 98,76% dos casos (−12,74% de EQM). O maior ganho, 14,99%, ocorre NOS DIAS DE VARIAÇÃO EXTREMA.",
         teste="Model Confidence Set",
         codigo="não",
         usar="Confirma o NOSSO efeito de cauda por via independente, em 404 ativos. E indica por que não superamos o HAR: eles usam variância intradiária de 5 min e 404 ativos; nós, Parkinson diário e 1 ativo. Justifica reformular a limitação como falta de PODER, não ausência de sinal.",
         ressalva="O período inclui 2020; convém verificar se o ganho sobrevive à exclusão da pandemia."),

    dict(rel="ALTA",
         autor="Mino, D.; Williamson, C.", ano="2025",
         titulo="Sentiment and Volatility in Financial Markets: BERT and GARCH during Geopolitical Crises",
         veiculo="arXiv:2510.16503",
         idioma="Inglês", mercado="S&P 500",
         volume="mais de 10.000 manchetes", periodo="jan–jul 2024 (105 dias)",
         tec="BERT ajustado ao domínio financeiro",
         alvo="volatilidade",
         modelo="GARCH(1,1) t-Student — o MESMO do nosso Script 04",
         resultado="Coeficiente do sentimento −0,2275 (p=0,0016). O nosso é −0,2924 (p=0,0002): magnitude equivalente, em outro mercado e idioma.",
         teste="testes t sobre os coeficientes",
         codigo="não",
         usar="Validação externa do nosso coeficiente. E ARGUMENTO DE DEFESA: eles NÃO avaliam fora da amostra e NÃO estratificam por regime (declaram como limitação). Nós fazemos as duas coisas. Eles param onde nós continuamos.",
         ressalva="Apenas 105 observações contra os nossos 1.988 pregões. Anuncia-se como revisão, mas traz análise empírica própria."),

    dict(rel="ALTA",
         autor="Kaplan, H. et al.", ano="2023",
         titulo="CrudeBERT: Applying Economic Theory towards Fine-Tuning Sentiment Models to the Crude Oil Market",
         veiculo="ICEIS 2023, p. 324–334 (arXiv:2305.06140)",
         idioma="Inglês", mercado="Mercado de petróleo bruto",
         volume="manchetes de choques de oferta e demanda", periodo="não declarado",
         tec="FinBERT ajustado; modelo em Captain-1337/CrudeBERT (777 downloads/mês)",
         alvo="movimentos do preço do petróleo",
         modelo="classificação de sentimento",
         resultado="Supera o FinBERT e classificadores tradicionais na previsão de movimentos do preço do petróleo.",
         teste="não recuperado",
         codigo="SIM — Hugging Face e GitHub (Captain-1337)",
         usar="A IDEIA, não o modelo. Rotular a manchete pelo MECANISMO ECONÔMICO (choque de oferta, de demanda, intervenção do controlador, política de preços, dividendos) em vez de positivo/negativo. RESPONDE À OBJEÇÃO DO PROF. EMERSON: desloca o critério do juízo subjetivo para a teoria econômica.",
         ressalva="Os rótulos de saída continuam sendo positivo/negativo/neutro; a inovação está na construção do conjunto de treinamento."),

    dict(rel="MEDIA",
         autor="Rahimikia, E.; Poon, S.-H.", ano="2021",
         titulo="Realised Volatility Forecasting: Machine Learning via Financial Word Embedding",
         veiculo="arXiv:2108.00480",
         idioma="Inglês", mercado="Ações dos EUA",
         volume="não recuperado", periodo="não recuperado",
         tec="embedding financeiro próprio (não usa cabeça de sentimento)",
         alvo="volatilidade realizada",
         modelo="aprendizado de máquina sobre HAR",
         resultado="Emprega embeddings em vez de classificação de sentimento para prever volatilidade.",
         teste="não recuperado",
         codigo="não verificado",
         usar="4ª confirmação independente da linha dos EMBEDDINGS. Todos os defeitos do FinBERT-PT-BR (viés de 87%, teto 0,58, sigmoide, zero pregões positivos) estão na CABEÇA DE CLASSIFICAÇÃO; nenhum afeta os embeddings.",
         ressalva="Ficha construída a partir de resultados de busca; convém ler o texto integral antes de citar."),

    # ------------------------------------------------- encoders de referência
    dict(rel="ALTA",
         autor="Araci, D. T.", ano="2019",
         titulo="FinBERT: Financial Sentiment Analysis with Pre-trained Language Models",
         veiculo="Dissertação de mestrado, Univ. de Amsterdã (arXiv:1908.10063)",
         idioma="Inglês", mercado="notícias financeiras em geral",
         volume="1,8 milhão de notícias (TRC2)", periodo="2008–2010",
         tec="bert-base-UNCASED + MLM financeiro + Financial PhraseBank",
         alvo="sentimento (3 classes)",
         modelo="classificação",
         resultado="4.459.091 downloads/mês; 778 citações (Semantic Scholar). Supera o estado da arte anterior.",
         teste="acurácia e F1 sobre conjuntos públicos",
         codigo="SIM — ProsusAI/finbert",
         usar="É `uncased`, portanto IMUNE ao nosso bug de caixa alta (Seção 4.j). Sugere a correção barata: normalizar tudo para minúsculas. E é PRECEDENTE: o artefato mais baixado da área nasceu como dissertação de mestrado.",
         ressalva="Corpus de 2008–2010: sujeito a deriva conceitual."),

    dict(rel="ALTA",
         autor="Huang, A. H.; Wang, H.; Yang, Y.", ano="2023",
         titulo="FinBERT: A Large Language Model for Extracting Information from Financial Text",
         veiculo="Contemporary Accounting Research, 40(2):806–841",
         idioma="Inglês", mercado="relatórios, teleconferências, analistas",
         volume="4,9 BILHÕES de tokens", periodo="—",
         tec="BERT + MLM; família finbert-tone, -esg, -fls, -pretrain",
         alvo="sentimento, ESG, declarações prospectivas",
         modelo="classificação",
         resultado="Supera o dicionário de Loughran-McDonald, naive Bayes, SVM, RF, CNN e LSTM. Vantagem justamente em sentenças que os demais rotulam como NEUTRAS. ESG: 89,5%.",
         teste="acurácia sobre conjuntos anotados",
         codigo="SIM — yiyanghkust/finbert-tone (704.839 downloads/mês)",
         usar="1) Referência revisada por pares, preferível ao preprint de Araci. 2) O modo de falha DELES resolvido é o NOSSO (90% dos nossos erros envolvem a classe Neutra) — evidência externa de que é problema de VOLUME de dados. 3) O modelo finbert-fls separa relato do passado de projeção de futuro: linha nova.",
         ressalva="Texto integral bloqueado (HTTP 403); números via fontes secundárias — conferir antes de citar."),

    dict(rel="MEDIA",
         autor="Shah, R. S. et al.", ano="2022",
         titulo="When FLUE Meets FLANG: Benchmarks and Large Pretrained Language Model for Financial Domain",
         veiculo="EMNLP 2022",
         idioma="Inglês", mercado="cinco tarefas financeiras",
         volume="—", periodo="—",
         tec="ELECTRA + MASCARAMENTO PREFERENCIAL de termos do domínio",
         alvo="sentimento, manchetes, NER, fronteiras, perguntas",
         modelo="—",
         resultado="Supera os modelos anteriores. FLUE é o primeiro benchmark aberto para linguagem financeira em inglês.",
         teste="benchmark multitarefa",
         codigo="SIM — SALT-NLP/FLANG",
         usar="Explica o fracasso do nosso experimento G3: usamos mascaramento ALEATÓRIO (Devlin), que eles mostram ser subótimo no domínio financeiro. Converte um resultado negativo em resultado com hipótese explicativa. E: NÃO EXISTE FLUE em português — lacuna de porte, candidata a doutorado.",
         ressalva="Reimplementar exigiria GPU e novo pré-treinamento. Registrar como trabalho futuro, não como ação imediata."),

    dict(rel="ALTA",
         autor="Malo, P. et al.", ano="2014",
         titulo="Good debt or bad debt: Detecting semantic orientations in economic texts (Financial PhraseBank)",
         veiculo="J. Assoc. Inf. Science and Technology, 65(4):782–796",
         idioma="Inglês", mercado="notícias e comunicados de empresas",
         volume="4.846 sentenças", periodo="—",
         tec="anotação humana — é o padrão-ouro sobre o qual os dois FinBERT são ajustados",
         alvo="sentimento (3 classes)",
         modelo="—",
         resultado="16 anotadores COM FORMAÇÃO EM FINANÇAS (3 pesquisadores + 13 mestrandos), 5 a 8 anotações POR SENTENÇA, 4 subconjuntos por concordância. Apenas 46,7% obtiveram acordo unânime.",
         teste="concordância entre anotadores",
         codigo="SIM — takala/financial_phrasebank",
         usar="RESPONDE À OBJEÇÃO DO PROF. EMERSON com protocolo citável. Ele tem razão sobre especialistas — mas a barra é 'mestrandos em finanças', alcançável na PUCPR. E o gargalo real não é a formação: é a REDUNDÂNCIA (5–8 anotações contra a nossa 1). Proposta: reanotar as MESMAS 300 com 3 anotadores.",
         ressalva="ATENÇÃO: o teste de teto mostrou que classificador perfeito rende só +1,2 p.p. na direção. Rotular melhor NÃO melhora a previsão; serve para saber se o 0,58 é culpa do modelo ou do anotador."),

    dict(rel="MEDIA",
         autor="Santos, L. L.; Bianchi, R. A. C.; Costa, A. H. R.", ano="2023",
         titulo="FinBERT-PT-BR: Análise de Sentimentos de Textos em Português do Mercado Financeiro",
         veiculo="BWAIF / SBC",
         idioma="Português", mercado="notícias financeiras brasileiras",
         volume="1,4 MILHÃO de textos; 503 rotulados", periodo="2006–2022",
         tec="BERTimbau (CASED) + MLM financeiro + ajuste fino",
         alvo="sentimento (3 classes)",
         modelo="classificação",
         resultado="76% de acurácia no domínio de origem; ~170 mil downloads/mês. É o modelo que a nossa pesquisa emprega.",
         teste="validação cruzada de 5 dobras",
         codigo="modelo SIM (lucas-leme/FinBERT-PT-BR); código de treino NÃO publicado",
         usar="É o nosso encoder. Auditado por nós pela primeira vez contra padrão humano (0,760 → 0,580 em subdomínio).",
         ressalva="ATENÇÃO: 1,4 milhão de TEXTOS (cartão do modelo), não 1,6 milhão de sentenças. O problem_type declarado como multi_label faz a pipeline aplicar SIGMOIDE. E é CASED, portanto vulnerável às 21.619 manchetes em caixa alta."),

    # ---------------------------------------- os números altos da tabela
    dict(rel="ATENCAO",
         autor="Bollen, J.; Mao, H.; Zeng, X.", ano="2011",
         titulo="Twitter mood predicts the stock market",
         veiculo="Journal of Computational Science, 2(1):1–8",
         idioma="Inglês", mercado="Índice DJIA (não é ação individual)",
         volume="9.853.498 tuítes", periodo="fev–dez 2008",
         tec="OpinionFinder + GPOMS (dicionários, 7 dimensões de humor)",
         alvo="direção do índice",
         modelo="rede neural difusa auto-organizável (SOFNN)",
         resultado="86,7% — que correspondem a 13 ACERTOS EM 15 PREGÕES (teste de 01 a 19/12/2008). Das 7 dimensões, SÓ 'calma' previu; positivo/negativo NÃO previu.",
         teste="Granger; sem correção para comparações múltiplas",
         codigo="não",
         usar="A ACURÁCIA não serve. Serve a descoberta esquecida: representação MULTIDIMENSIONAL de humor supera positivo/negativo. Casa com o nosso erro concentrado no Neutro (90%) e com Pos×Neg dando 0,783.",
         ressalva="REFUTADO. Lachanski e Pav (2017), Econ Journal Watch: estenderam a série para 2007 e NÃO acharam evidência fora da amostra; atribuíram a data snooping e viés de comparações múltiplas."),

    dict(rel="ATENCAO",
         autor="Schumaker, R. P.; Chen, H.", ano="2009",
         titulo="A quantitative stock prediction system based on financial news (AZFinText)",
         veiculo="Information Processing & Management, 45:571–583",
         idioma="Inglês", mercado="Ações do S&P 500",
         volume="9.211 notícias; 10,2 milhões de cotações", periodo="5 semanas",
         tec="saco de palavras, sintagmas nominais, entidades nomeadas",
         alvo="preço 20 MINUTOS após a notícia (não é o dia seguinte)",
         modelo="derivado de SVM para predição numérica discreta",
         resultado="71,18% de acurácia direcional — o MELHOR entre vários esquemas de particionamento; retorno simulado de 8,50%.",
         teste="não declarado no resumo",
         codigo="não",
         usar="A ACURÁCIA não é comparável: mede REAÇÃO em 20 min, que é o nosso horizonte P0, e não previsão do dia seguinte (P1). Serve a LIÇÃO: o sinal vive no curtíssimo prazo — 3º apoio para buscar dados INTRADIÁRIOS da PETR4.",
         ressalva="Cinco semanas de dados. É o melhor entre esquemas. Comparar com os nossos 54,5% seria erro de leitura."),

    dict(rel="ATENCAO",
         autor="Barak, S.; Arjmand, A.; Ortobelli, S.", ano="2017",
         titulo="Fusion of multiple diverse predictors in stock market",
         veiculo="Information Fusion",
         idioma="—", mercado="Bolsa de Teerã (mercado menos líquido)",
         volume="—", periodo="—",
         tec="ensembles: bagging, boosting, AdaBoost + meta-classificador",
         alvo="retorno e risco (não direção diária a partir de notícias)",
         modelo="fusão de preditores diversos",
         resultado="Até 83,6% (retorno) e 88,2% (risco) — o MÁXIMO entre várias configurações.",
         teste="não declarado",
         codigo="não",
         usar="NADA A APROVEITAR: a técnica JÁ FOI REPLICADA na Seção 4.d. O stacking rendeu 53,14% / 52,99% / 53,14%, todos no baseline de classe majoritária ou abaixo. O XGBoost simples com 3 atributos deu 54,52% e superou todos.",
         ressalva="Mercado e tarefa distintos. Reporta a acurácia máxima entre configurações."),

    dict(rel="ALTA",
         autor="Lachanski, M.; Pav, S.", ano="2017",
         titulo="Shy of the Character Limit: 'Twitter Mood Predicts the Stock Market' Revisited",
         veiculo="Econ Journal Watch, 14(3):302–345",
         idioma="Inglês", mercado="Índice DJIA",
         volume="reconstrução da série de Bollen", periodo="2007–2008",
         tec="replicação",
         alvo="—",
         modelo="—",
         resultado="NÃO encontraram evidência de que o humor do Twitter auxilie a previsão fora da amostra. Atribuem o achado original a data snooping e a viés de comparações múltiplas.",
         teste="replicação com amostra estendida",
         codigo="não",
         usar="MUNIÇÃO DE DEFESA. Se a banca citar os 86,7% de Bollen, responder com esta referência. Já incorporada à Seção 4 e ao bib (lachanski_shy_2017).",
         ressalva="—"),

    # ------------------------------------- citantes do FinBERT-PT-BR (PT-BR)
    dict(rel="ALTA",
         autor="Pinheiro, T.; Muinhos, M. K.; Fernandes, M.", ano="2025",
         titulo="The role of fiscal sentiment in Brazil's yield curve",
         veiculo="artigo em congresso (PDF público)",
         idioma="Português", mercado="Curva de juros brasileira",
         volume="jornais + Broadcast da Agência Estado", periodo="2008–2022",
         tec="FinBERT-PT-BR como extrator de EMBEDDINGS + K-means++",
         alvo="estrutura a termo da curva de juros",
         modelo="modelo macrofinanceiro de estrutura a termo",
         resultado="Constroem o Índice de Sentimento Fiscal (FSI-BR) de alta frequência. Sentimento no nível da SENTENÇA, remapeado para {−1, 0, +1}. Dicionários fiscais ENDÓGENOS por governo.",
         teste="—",
         codigo="não",
         usar="Usam o FinBERT-PT-BR como EXTRATOR DE EMBEDDINGS, não como classificador — contorna todos os defeitos da cabeça de sentimento. E a extração no nível da SENTENÇA (não do documento) é refinamento que não fazemos.",
         ressalva="Fonte encontrada via NotebookLM e verificada por busca. Ficha ainda não construída em detalhe."),

    dict(rel="MEDIA",
         autor="Costa Neto, A. M.; Anjos, L. C. M.", ano="2025",
         titulo="Informatividade de notas explicativas (Boilerplateness, Completeness, Density)",
         veiculo="Congresso USP/FIPECAFI",
         idioma="Português", mercado="1.152 empresas listadas na CVM",
         volume="25.804 notas explicativas", periodo="2011–2023",
         tec="FinBERT-PT-BR como extrator de EMBEDDINGS + K-means",
         alvo="índice de informatividade contábil",
         modelo="agrupamento sobre embeddings; Gradient Boosting",
         resultado="Três dimensões: Boilerplateness (repetição sem valor), Completeness (cobertura) e Density (densidade semântica).",
         teste="—",
         codigo="não",
         usar="Segunda confirmação da linha dos EMBEDDINGS em português. A ideia de medir REPETIÇÃO (boilerplate) é adaptável: manchetes repetidas entre portais inflam o nosso índice diário.",
         ressalva="Domínio contábil, não de mercado. Verificar a referência completa antes de citar."),

    dict(rel="MEDIA",
         autor="Błoch, A.; Santana, R.; Amantino, M.", ano="2026",
         titulo="Os jesuítas e a Era do Algoritmo: análise de sentimentos da correspondência colonial",
         veiculo="Estudos Ibero-Americanos",
         idioma="Português", mercado="correspondência colonial (não financeiro)",
         volume="registros de 1642 a 1822", periodo="1642–1822",
         tec="FinBERT-PT-BR em comitê (máquina de comitê)",
         alvo="sentimento em textos históricos",
         modelo="comitê de modelos",
         resultado="É o ÚNICO trabalho localizado que efetivamente EXECUTOU o FinBERT-PT-BR em comitê.",
         teste="—",
         codigo="não",
         usar="Método de comitê replicável — JÁ TESTADO por nós, e PIOROU o desempenho.",
         ressalva="Domínio completamente distinto (textos históricos)."),

    dict(rel="MEDIA",
         autor="Imai, R. et al.", ano="2024",
         titulo="Is it Fine to Tune? Evaluating SentenceBERT Fine-tuning for Brazilian Portuguese Text Streams",
         veiculo="PPGIa/PUCPR",
         idioma="Português", mercado="notícias de portais de economia",
         volume="fluxo contínuo de notícias", periodo="—",
         tec="SBERTimbau + Adaptive Random Forest",
         alvo="classificação em editorias",
         modelo="aprendizado em fluxo, sensível a deriva",
         resultado="Ajuste fino incremental ANUAL (amostra de ~2.000 notícias do ano anterior) mantém o F1 estável ao longo do tempo.",
         teste="avaliação em fluxo",
         codigo="não",
         usar="AMEAÇA DIRETA À NOSSA VALIDADE: deriva conceitual ao longo de 2018–2025. E a receita deles (reajuste incremental anual) é adaptável. Autores do nosso próprio programa.",
         ressalva="Tarefa distinta (editoria, não sentimento)."),

    dict(rel="MEDIA",
         autor="Teles, A.; Figueiredo, C.", ano="2025",
         titulo="Comparing LLMs for Sentiment Analysis in Financial Market News",
         veiculo="arXiv:2510.15929",
         idioma="Português", mercado="notícias do mercado financeiro",
         volume="—", periodo="—",
         tec="LLMs generativos contra classificadores clássicos",
         alvo="sentimento",
         modelo="comparação",
         resultado="Compara a precisão de diferentes LLMs na interpretação de notícias financeiras.",
         teste="—",
         codigo="não",
         usar="Definiu o nosso experimento G6, em que o Qwen2.5-3B obteve 0,480 contra 0,580 do FinBERT-PT-BR. Confirmado externamente: em ESG, o FinBERT com 83% superou TODOS os LLMs testados.",
         ressalva="—"),

    dict(rel="MEDIA",
         autor="Reichert, B.; Perlin, M.", ano="2025",
         titulo="Dicionário financeiro em português gerado por LLM",
         veiculo="Computational Economics",
         idioma="Português", mercado="Fatos Relevantes e comunicados do COPOM",
         volume="—", periodo="—",
         tec="dicionário gerado por ChatGPT; usa o SentFinBERT-PT-BR como referência",
         alvo="sentimento binário",
         modelo="soma de escores de palavras",
         resultado="Usam o FinBERT-PT-BR como PADRÃO-OURO de contexto completo. Registram o custo: o transformer leva mais de uma hora para inferir grandes volumes; o dicionário é instantâneo.",
         teste="acurácia contra o transformer",
         codigo="não",
         usar="Linha de base LEXICAL que nos falta. Comparar o nosso índice contextual contra um dicionário é exigência natural da banca.",
         ressalva="—"),

    dict(rel="ALTA",
         autor="Silva, C. G.", ano="2018",
         titulo="Efeito do sentimento de notícias sobre retorno e volatilidade do IBOVESPA",
         veiculo="Tese de doutorado",
         idioma="Português", mercado="Índice IBOVESPA",
         volume="—", periodo="—",
         tec="análise de sentimento; regressão quantílica",
         alvo="retorno e volatilidade",
         modelo="GARCH; regressão quantílica com pesos variáveis",
         resultado="O sentimento afeta retorno e volatilidade. Sentimento linear isolado: R² fora da amostra NEGATIVO. O ganho vem da combinação quantílica.",
         teste="R² fora da amostra",
         codigo="não",
         usar="Base metodológica já replicada: a quantílica ponderada rendeu +10,9% de R² fora da amostra. A nossa contribuição é ESTENDER do índice agregado para um ATIVO INDIVIDUAL.",
         ressalva="Índice agregado, não ação individual — a comparação direta de magnitudes não é válida."),

    dict(rel="MEDIA",
         autor="Nguyen, T. H.; Shirai, K.; Velcin, J.", ano="2015",
         titulo="Sentiment analysis on social media for stock movement prediction",
         veiculo="Expert Systems with Applications",
         idioma="Inglês", mercado="Ações dos EUA",
         volume="—", periodo="—",
         tec="TSLDA — sentimento por TÓPICO",
         alvo="direção de ações",
         modelo="—",
         resultado="Ganho de 2,1 a 9,8 pontos percentuais sobre a linha de base de preços.",
         teste="—",
         codigo="não",
         usar="É a comparação HONESTA para nós: não a acurácia absoluta, mas o GANHO INCREMENTAL do sentimento sobre preços. O nosso ganho de 4,4 p.p. está dentro dessa faixa.",
         ressalva="Sentimento por tópico já replicado (atributos por categoria) e PIOROU fora da amostra."),

    dict(rel="BAIXA",
         autor="Li, X. et al.", ano="2020",
         titulo="Incorporating stock prices and news sentiments for stock movement prediction",
         veiculo="Information Processing & Management",
         idioma="Inglês", mercado="Hong Kong",
         volume="—", periodo="—",
         tec="fusão sequencial em aprendizado profundo",
         alvo="tendência",
         modelo="deep learning",
         resultado="Supera as linhas de base ao fundir preços e sentimento.",
         teste="—",
         codigo="não",
         usar="Corrobora a arquitetura de fusão que adotamos. Fusão sequencial profunda fica como trabalho futuro.",
         ressalva="Resultado reportado sem magnitude precisa na nossa tabela."),
]

CORES = {"NOSSA": NOSSO, "ALTA": VERDE, "MEDIA": AMARELO,
         "BAIXA": CINZA, "ATENCAO": VERMELHO}
ROTULO_REL = {"NOSSA": "A NOSSA", "ALTA": "ALTA", "MEDIA": "Média",
              "BAIXA": "Baixa", "ATENCAO": "ATENÇÃO"}


def _cabecalho(ws, colunas, titulo, subtitulo):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(colunas))
    c = ws.cell(row=1, column=1, value=titulo)
    c.font = Font(bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=AZUL)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(colunas))
    c = ws.cell(row=2, column=2 - 1, value=subtitulo)
    c.font = Font(italic=True, size=9, color="404040")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    for j, (nome, larg) in enumerate(colunas, start=1):
        c = ws.cell(row=3, column=j, value=nome)
        c.font = Font(bold=True, size=10, color=AZUL)
        c.fill = PatternFill("solid", fgColor=CINZA_CAB)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDA
        ws.column_dimensions[get_column_letter(j)].width = larg
    ws.row_dimensions[3].height = 34


def aba_comparativo(wb):
    ws = wb.active
    ws.title = "Comparativo"
    _cabecalho(ws, COLUNAS,
               "PESQUISAS QUE LEEM NOTÍCIAS PARA PREVER DIREÇÃO E VOLATILIDADE",
               "Levantamento para a mentoria com o Prof. Dr. Emerson Cabrera Paraiso · "
               "Vanderlei Barbosa da Silva · PUCPR/PPGIa · agosto de 2026")

    ordem = ["autor", "ano", "titulo", "veiculo", "idioma", "mercado", "volume",
             "periodo", "tec", "alvo", "modelo", "resultado", "teste", "codigo",
             "rel", "usar", "ressalva"]

    for i, linha in enumerate(LINHAS, start=1):
        r = 3 + i
        cor = CORES[linha["rel"]]
        ws.cell(row=r, column=1, value=i)
        for j, chave in enumerate(ordem, start=2):
            v = ROTULO_REL[linha[chave]] if chave == "rel" else linha[chave]
            ws.cell(row=r, column=j, value=v)
        for j in range(1, len(COLUNAS) + 1):
            c = ws.cell(row=r, column=j)
            c.fill = PatternFill("solid", fgColor=cor)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = BORDA
            c.font = Font(size=9, bold=(linha["rel"] in ("NOSSA", "ATENCAO")
                                        and j in (2, 16)))
        ws.row_dimensions[r].height = 92

    ws.freeze_panes = "C4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(COLUNAS))}{3 + len(LINHAS)}"

    # legenda
    r = 3 + len(LINHAS) + 2
    ws.cell(row=r, column=1, value="LEGENDA DAS CORES").font = Font(bold=True, size=10)
    for k, (chave, texto) in enumerate([
        ("NOSSA", "A nossa pesquisa"),
        ("ALTA", "Relevância ALTA — há algo concreto a aproveitar"),
        ("MEDIA", "Relevância média — contribui de forma indireta"),
        ("BAIXA", "Relevância baixa — apenas contextual"),
        ("ATENCAO", "ATENÇÃO — número não comparável, refutado ou já testado sem êxito"),
    ], start=1):
        c = ws.cell(row=r + k, column=1, value="")
        c.fill = PatternFill("solid", fgColor=CORES[chave])
        c.border = BORDA
        ws.cell(row=r + k, column=2, value=texto).font = Font(size=9)


def aba_nossos_numeros(wb):
    cols = [("Item", 40), ("Valor", 22), ("Significância", 22),
            ("Onde está", 22), ("Observação", 62)]
    ws = wb.create_sheet("Nossos números")
    _cabecalho(ws, cols, "OS NÚMEROS DA NOSSA PESQUISA, JÁ CORRIGIDOS",
               "Use esta aba como referência — inclui as correções de agosto de 2026")

    dados = [
        ("Direção — XGBoost, 3 atributos", "54,5%", "binomial p=0,012",
         "Seção 4.d", "Ganho de 4,4 p.p. sobre o modelo apenas-preços. É ESTE o número a usar."),
        ("Direção — ponderação por confiança", "50,31% (teste)", "—",
         "Seção 4.n", "ATENÇÃO: os 54,93% que circulam são acurácia de VALIDAÇÃO. No teste a ponderação PIORA (53,88% sem ponderar)."),
        ("Direção — regra ingênua sobre o rótulo", "47,59%", "p<0,0001",
         "Seção 4.l", "Abaixo do mercado (52,78%). Efeito do viés de negatividade contra a tendência de alta."),
        ("Classificação — acurácia contra humano", "0,580", "kappa 0,371",
         "Seção 4.g", "O FinBERT INGLÊS obtém 0,555 em situação equivalente. O nosso patamar é NORMAL."),
        ("Coeficiente do sentimento na volatilidade", "−0,2924", "p=0,0002",
         "Seção 4.k", "Mino e Williamson (2025) obtêm −0,2275 no S&P 500. Magnitude equivalente."),
        ("Filtro de relevância (associação)", "|r| 0,1385 → 0,1704", "p=0,0010",
         "Seção 4.k", "+23%. Único ganho em nove intervenções. Recorte CAT1+CAT2."),
        ("Previsão de volatilidade contra o HAR", "não supera", "p=0,6405",
         "Seção 4.k", "Compatível com falta de PODER (medida diária, ativo único), não com ausência de sinal."),
        ("Notícia da EMPRESA, horizonte de 22 dias", "+1,77% de EQM", "p=0,0574",
         "Seção 4.o", "MELHOR RESULTADO DA PESQUISA até agora. Ficou pouco aquém do limiar de 5%."),
        ("Notícia MACRO, horizontes de 5 e 22 dias", "−1,09% e −1,79%", "p=0,0146 e p=0,0200",
         "Seção 4.o", "PIORA de forma significativa. Inverte a hipótese de Bodilsen e Lunde (2025)."),
        ("Efeito de cauda — Pearson", "−0,1309", "p<0,0001",
         "Seção 4.l", "Sensível às magnitudes: detecta a associação."),
        ("Efeito de cauda — Spearman", "−0,0268", "p=0,2367",
         "Seção 4.l", "Só ordenação: NÃO detecta. Prova que o efeito vive nos extremos."),
        ("Regressão quantílica, quantil 0,05", "+542 pontos-base", "p=0,001",
         "Seção 4", "Nulo nos quantis altos. Converge com o efeito de cauda por via independente."),
        ("Pregões com maioria positiva", "ZERO de 1.989", "—",
         "Seção 4.l", "ISM negativo em 100% dos dias. Limitação declarada no Capítulo 5."),
        ("Adaptação de domínio (G3)", "F1 −0,056", "p=0,022",
         "Seção 4.i", "Perplexidade caiu 49% mas a classificação DEGRADOU. Com 352 exemplos; a literatura usa 1.500."),
    ]
    for i, linha in enumerate(dados, start=1):
        r = 3 + i
        for j, v in enumerate(linha, start=1):
            c = ws.cell(row=r, column=j, value=v)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = BORDA
            c.font = Font(size=9)
            if "ATENÇÃO" in str(linha[4]) or "MELHOR" in str(linha[4]):
                c.fill = PatternFill("solid", fgColor=VERMELHO if "ATENÇÃO" in str(linha[4]) else VERDE)
        ws.row_dimensions[r].height = 46
    ws.freeze_panes = "A4"


def aba_como_usar(wb):
    cols = [("Prioridade", 11), ("Ação proposta", 40), ("De onde veio", 28),
            ("Por que deve funcionar", 60), ("Custo", 14), ("Depende de", 26)]
    ws = wb.create_sheet("Como usar")
    _cabecalho(ws, cols, "PLANO DE ADAPTAÇÃO — O QUE FAZER COM O QUE FOI ENCONTRADO",
               "Ordenado por razão entre ganho esperado e custo")

    dados = [
        ("1", "Mudar o alvo para a DIREÇÃO da volatilidade (amanhã sacode mais ou menos que hoje?)",
         "Hashamia e Maldonado (2025)",
         "Via do meio: a direção do preço é quase acaso e o nível da volatilidade o HAR já prevê bem. O HAR não é adversário natural da MUDANÇA de nível. Nunca testado aqui.",
         "BAIXO", "nada — os dados já existem"),
        ("2", "Adotar o recorte EMPRESA e varrer horizontes entre 10 e 30 pregões",
         "Nosso experimento (Seção 4.o)",
         "Já rendeu +1,77% com p=0,0574 em 22 dias — o melhor resultado da pesquisa. Um horizonte intermediário pode cruzar o limiar de 5%.",
         "MUITO BAIXO", "rodar o script com outros valores"),
        ("3", "Testar a CONTAGEM de notícias contra a direção da volatilidade",
         "Hashamia e Maldonado (2025)",
         "A contagem superou o sentimento no estudo deles. Aqui falhou contra o NÍVEL (p=0,222), mas nunca foi testada contra a DIREÇÃO.",
         "BAIXO", "nada"),
        ("4", "Normalizar as manchetes para minúsculas antes de classificar",
         "Araci (2019) — FinBERT é uncased",
         "As 21.619 manchetes em caixa alta degradam o nosso modelo cased (cobertura de vocabulário 22% contra 79%). O modelo inglês é imune por decisão de projeto.",
         "BAIXO", "reprocessar o corpus (GPU)"),
        ("5", "Usar EMBEDDINGS em vez da cabeça de sentimento",
         "4 fontes independentes",
         "TODOS os defeitos documentados (viés de 87%, teto 0,58, sigmoide, zero pregões positivos) estão na cabeça de classificação. NENHUM afeta os embeddings, que vêm da parte treinada com 1,4 milhão de textos.",
         "MÉDIO", "GPU (Colab)"),
        ("6", "Rotular pelo MECANISMO ECONÔMICO (oferta, demanda, intervenção, preços, dividendos)",
         "CrudeBERT (Kaplan et al., 2023)",
         "Substitui juízo subjetivo por classificação de fato. RESPONDE À OBJEÇÃO DO PROF. EMERSON sobre a necessidade de especialistas, deslocando o critério para a teoria econômica.",
         "MÉDIO", "rotulagem, mas com critério objetivo"),
        ("7", "Extrair sentimento no nível da SENTENÇA, não do documento",
         "Pinheiro, Muinhos e Fernandes (2025)",
         "Hoje concatenamos título e resumo num único texto. A extração por sentença é mais granular e é o que fazem no FSI-BR.",
         "MÉDIO", "reprocessar o corpus (GPU)"),
        ("8", "Reanotar as MESMAS 300 manchetes com 3 anotadores",
         "Malo et al. (2014)",
         "O padrão-ouro internacional usa 5 a 8 anotações por sentença. Sem redundância não há como calcular concordância, e sem concordância não se distingue erro do modelo de ruído do anotador.",
         "MÉDIO", "3 anotadores com formação em finanças"),
        ("9", "Comparar contra uma linha de base LEXICAL (dicionário)",
         "Reichert e Perlin (2025)",
         "A banca perguntará se o transformer supera um dicionário simples. Hoje não temos essa comparação.",
         "BAIXO", "dicionário público em português"),
        ("10", "Obter dados INTRADIÁRIOS da PETR4",
         "3 apoios independentes",
         "Schumaker (20 min), o nosso contraste P0 contra P1, e Halousková e Lyócsa (5 min, que superam o HAR). O sinal vive no curtíssimo prazo.",
         "ALTO", "fonte de dados intradiários"),
        ("11", "Replicar o pipeline para 5 a 10 ativos líquidos da B3",
         "Halousková e Lyócsa (2025)",
         "Eles usam 404 ativos. Com um só, falta poder estatístico. Permite ainda testar a predição de que o efeito de cauda cresce com a volatilidade do ativo.",
         "MÉDIO", "coleta para outros ativos"),
        ("12", "Testar mascaramento PREFERENCIAL de termos do domínio",
         "Shah et al. (2022) — FLANG",
         "Explica o fracasso do G3: usamos mascaramento aleatório, que eles mostram ser subótimo em finanças. Temos a taxonomia de 152 termos pronta.",
         "ALTO", "GPU e novo pré-treinamento"),
    ]
    for i, linha in enumerate(dados, start=1):
        r = 3 + i
        for j, v in enumerate(linha, start=1):
            c = ws.cell(row=r, column=j, value=v)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = BORDA
            c.font = Font(size=9, bold=(j == 1))
            if linha[4] in ("MUITO BAIXO", "BAIXO"):
                c.fill = PatternFill("solid", fgColor=VERDE)
            elif linha[4] == "MÉDIO":
                c.fill = PatternFill("solid", fgColor=AMARELO)
            else:
                c.fill = PatternFill("solid", fgColor=CINZA)
        ws.row_dimensions[r].height = 62
    ws.freeze_panes = "A4"


def main() -> None:
    wb = Workbook()
    aba_comparativo(wb)
    aba_nossos_numeros(wb)
    aba_como_usar(wb)
    wb.save(SAIDA)
    print(f"[OK] Planilha gerada: {SAIDA}")
    print(f"     Aba 1 'Comparativo'    : {len(LINHAS)} pesquisas x {len(COLUNAS)} colunas")
    print(f"     Aba 2 'Nossos numeros' : os nossos resultados, ja corrigidos")
    print(f"     Aba 3 'Como usar'      : 12 acoes por prioridade")


if __name__ == "__main__":
    main()
