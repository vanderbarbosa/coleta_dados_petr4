# -*- coding: utf-8 -*-
# ==============================================================================
#   Planilha das seis essenciais — 3 direção, 3 volatilidade
#   Saída: Mentoria_Emerson_13082026/09_AS_SEIS_ESSENCIAIS.xlsx
# ==============================================================================
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

AQUI = Path(__file__).resolve().parent
SAIDA = AQUI / "09_AS_SEIS_ESSENCIAIS.xlsx"

AZUL, CAB = "1F3864", "D9E2F3"
VOL, DIR, NOSSO, DEST = "DDEBF7", "E2EFDA", "FFE699", "FCE4E4"
B = Border(*[Side(style="thin", color="BFBFBF")] * 4)

COLS = [("#", 4), ("Pesquisa", 24), ("Ano", 6), ("Veículo", 24),
        ("Alvo — o que prevê", 26), ("Mercado / Ativo", 20),
        ("Corpus / Volume", 22), ("Período", 13),
        ("Encoder / Tecnologia", 28), ("Referência (baseline)", 18),
        ("MELHOR RESULTADO", 46), ("Significância", 18),
        ("PDF?", 9), ("Por que importa para nós", 50),
        ("O que fazer com isso", 40)]

SEC_VOL = ("VOLATILIDADE — as três melhores e mais próximas",)
SEC_DIR = ("DIREÇÃO DO PREÇO — as três melhores e mais próximas",)

L = [
 ("H", SEC_VOL[0]),

 ("V", "Halousková e Lyócsa", "2025", "arXiv:2503.19767",
  "volatilidade (variância realizada)", "404 ações do S&P 500",
  "jornais, Twitter, Google Trends, Wikipédia", "2010–2021",
  "FinBERT (Araci) + HAR; subconjuntos completos e LASSO adaptativo",
  "HAR",
  "VENCE o HAR em 98,76% das ações. −12,74% de erro médio; −14,99% NOS DIAS EXTREMOS.",
  "Model Confidence Set", "SIM",
  "Confirma o nosso EFEITO DE CAUDA por via independente, em 404 ativos. E nos supera — mas com 404 ações e dado de 5 minutos, contra 1 ação e dado diário.",
  "Justificar a limitação como falta de PODER, não ausência de sinal. Replicar para 5–10 ativos da B3."),

 ("D", "Hashami e Maldonado", "2025", "arXiv:2508.20707",
  "DIREÇÃO da volatilidade (binária)", "Futuros de Brent (PETRÓLEO)",
  "notícias Eikon; ~592 mil manchetes", "2014–2024",
  "VADER, TextBlob, FinBERT, CrudeBERT; embeddings GloVe, FastText, BERT, Gemini, LLaMA",
  "HAR (0,6494)",
  "FastText 0,7136 | contagem de notícias 0,7054 | FinBERT-embedding 0,6694 | FinBERT-sentimento 0,5368",
  "McNemar", "SIM",
  "A MAIS IMPORTANTE. Petróleo é o nosso ativo; código público. E traz a prova controlada: o MESMO FinBERT dá 0,5368 na cabeça de sentimento e 0,6694 como embedding.",
  "1) adotar o alvo direção da volatilidade; 2) testar contagem de notícias; 3) extrair embeddings do FinBERT-PT-BR"),

 ("V", "Bodilsen e Lunde", "2025", "J. Applied Econometrics 40(1):18-36",
  "volatilidade realizada, vários horizontes", "Ações dos EUA e S&P 500",
  "analítica de notícias comercial", "não recuperado",
  "analítica comercial + família HAR",
  "HAR",
  "Notícia MACRO melhora significativamente; notícia da EMPRESA não acrescenta. Ganho maior em HORIZONTE LONGO.",
  "comparação preditiva", "não (Wiley)",
  "Periódico de primeira linha. Foi ao testá-la que obtivemos o MELHOR RESULTADO da pesquisa. Na PETR4 a conclusão INVERTE-SE: empresa ajuda, macro atrapalha.",
  "Adotar recorte EMPRESA; varrer horizontes de 10 a 30 pregões (hoje: +1,77% em 22d, p=0,0574)"),

 ("H", SEC_DIR[0]),

 ("D", "Ruan e Jiang", "2025", "Mathematics 13(17):2747",
  "DIREÇÃO do preço", "Ações do S&P 500",
  "manchetes financeiras", "2018–2023",
  "FinBERT + XGBoost sobre preço e volatilidade; SHAP; privacidade diferencial",
  "técnica-apenas e léxica",
  "Supera as bases em AUC, F1 e lucro simulado. SHAP: sentimento pesa 28,6%; volatilidade 21,4%.",
  "não recuperado", "não (MDPI)",
  "ARQUITETURA QUASE IDÊNTICA À NOSSA: FinBERT + preço + volatilidade em XGBoost, mesmo período. É a prova de que o nosso desenho está correto.",
  "Acrescentar SHAP — é biblioteca pronta, custa horas, e a banca valoriza"),

 ("D", "Nguyen, Shirai e Velcin", "2015", "Expert Systems with Applications",
  "DIREÇÃO de ações", "Ações dos EUA",
  "notícias e redes sociais", "não recuperado",
  "TSLDA — sentimento por TÓPICO",
  "modelo só-preços",
  "GANHO de 2,1 a 9,8 PONTOS PERCENTUAIS sobre a base só-preços.",
  "não recuperado", "não (Elsevier)",
  "É A NOSSA RÉGUA. A comparação justa entre estudos não é a acurácia absoluta, e sim o ganho da notícia. O NOSSO GANHO É 4,4 p.p. — dentro da faixa.",
  "Apresentar SEMPRE os 4,4 p.p., nunca os 54,5% isolados"),

 ("D", "Schumaker e Chen", "2009", "Inf. Processing & Management 45:571-583",
  "preço 20 MINUTOS após a notícia", "Ações do S&P 500",
  "9.211 notícias; 10,2 mi de cotações", "5 semanas",
  "saco de palavras, sintagmas nominais, entidades + SVM",
  "—",
  "71,18% de acurácia direcional; retorno simulado de 8,50%. É o MELHOR entre vários esquemas.",
  "não declarado", "não (Elsevier)",
  "O maior número LEGÍTIMO de direção do conjunto. Mas mede REAÇÃO em 20 min — é o nosso horizonte P0, não o P1. Nós já medimos o mesmo padrão: P0 ordenado (55,0/53,2/51,6), P1 colapsa.",
  "3º apoio para buscar dados INTRADIÁRIOS da PETR4"),

 ("H", "A NOSSA PESQUISA, para comparação"),

 ("N", "Silva (a nossa)", "2026", "Dissertação PUCPR/PPGIa",
  "DIREÇÃO e VOLATILIDADE (as duas)", "B3 — PETR4 (ação individual)",
  "205.697 manchetes de 5 portais", "2018–2025",
  "FinBERT-PT-BR + XGBoost/SVM + GARCH(1,1)-t + HAR + quantílica",
  "só-preços e HAR",
  "Direção 54,5% (GANHO DE 4,4 p.p., p=0,012). Volatilidade: coef. −0,2924 (p=0,0002) mas NÃO supera o HAR (p=0,64). Melhor: empresa em 22d, +1,77% (p=0,0574).",
  "binomial, McNemar, Diebold-Mariano, bootstrap", "—",
  "O que temos e nenhum deles tem: AUDITORIA do artefato — viés de 48,5% negativos, escore em escala errada, caixa alta, ZERO pregões de maioria positiva em 8 anos.",
  "—"),
]

CORES = {"V": VOL, "D": DEST, "N": NOSSO}


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "As seis"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
    c = ws.cell(row=1, column=1,
                value="AS SEIS PESQUISAS ESSENCIAIS — 3 DE VOLATILIDADE, 3 DE DIREÇÃO")
    c.font = Font(bold=True, size=13, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=AZUL)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLS))
    c = ws.cell(row=2, column=1,
                value="Selecionadas por proximidade do nosso desenho cruzada com a qualidade "
                      "do resultado · Vanderlei Barbosa da Silva · PUCPR/PPGIa · agosto de 2026")
    c.font = Font(italic=True, size=9, color="404040")
    c.alignment = Alignment(horizontal="center")

    for j, (nome, larg) in enumerate(COLS, start=1):
        c = ws.cell(row=3, column=j, value=nome)
        c.font = Font(bold=True, size=10, color=AZUL)
        c.fill = PatternFill("solid", fgColor=CAB)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = B
        ws.column_dimensions[get_column_letter(j)].width = larg
    ws.row_dimensions[3].height = 34

    n = 0
    for i, item in enumerate(L):
        r = 4 + i
        if item[0] == "H":
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(COLS))
            c = ws.cell(row=r, column=1, value=item[1])
            c.font = Font(bold=True, size=11, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=AZUL)
            c.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[r].height = 20
            continue
        n += 1
        ws.cell(row=r, column=1, value=n)
        for j, v in enumerate(item[1:], start=2):
            ws.cell(row=r, column=j, value=v)
        for j in range(1, len(COLS) + 1):
            c = ws.cell(row=r, column=j)
            c.fill = PatternFill("solid", fgColor=CORES[item[0]])
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = B
            c.font = Font(size=9, bold=(j in (2, 11)))
        ws.row_dimensions[r].height = 108

    ws.freeze_panes = "C4"

    r = 4 + len(L) + 1
    for k, txt in enumerate([
        "AS SEIS LIÇÕES, EM UMA LINHA CADA:",
        "1. O efeito é de CAUDA — confirmado por Halousková em 404 ações.",
        "2. EMBEDDING vence cabeça de sentimento — o MESMO FinBERT: 0,5368 contra 0,6694 (Hashami).",
        "3. CONTAR notícias vence medir o tom — 0,7054, melhor que todo método de sentimento (Hashami).",
        "4. PRAZO LONGO é melhor que um dia — Bodilsen, e confirmado no nosso teste (+1,77% em 22 dias).",
        "5. A NOSSA ARQUITETURA está correta — Ruan e Jiang usam a mesma. Falta-nos o SHAP.",
        "6. A RÉGUA é o GANHO, não a acurácia — a literatura dá 2 a 10 p.p.; o nosso é 4,4.",
    ]):
        c = ws.cell(row=r + k, column=1, value=txt)
        c.font = Font(bold=(k == 0), size=10)
        ws.merge_cells(start_row=r + k, start_column=1, end_row=r + k, end_column=len(COLS))

    wb.save(SAIDA)
    print(f"[OK] {SAIDA}  ({n} linhas)")


if __name__ == "__main__":
    main()
