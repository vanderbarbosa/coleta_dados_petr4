# -*- coding: utf-8 -*-
# ==============================================================================
#   Planilha focada: SOMENTE pesquisas que preveem direção e/ou volatilidade
#   Saída: Mentoria_Emerson_13082026/05_PREVISAO_DIRECAO_E_VOLATILIDADE.xlsx
# ==============================================================================
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

AQUI = Path(__file__).resolve().parent
SAIDA = AQUI / "05_PREVISAO_DIRECAO_E_VOLATILIDADE.xlsx"

AZUL, CAB = "1F3864", "D9E2F3"
VERDE, AMARELO, VERMELHO, NOSSO = "E2EFDA", "FFF2CC", "FCE4E4", "FFE699"
B = Border(*[Side(style="thin", color="BFBFBF")] * 4)

COLS = [("#", 4), ("Autor / Ano", 22), ("Alvo — o que tenta prever", 30),
        ("Mercado / Ativo", 22), ("Encoder / Tecnologia", 26),
        ("Volume de notícias", 17), ("Resultado", 42),
        ("Bate a referência?", 17), ("Vale comparar com o nosso?", 34)]

# cor: N=nosso  V=vitoria deles  A=atencao (nao comparavel)  M=neutro
L = [
 # ---------------------------------------------------------------- VOLATILIDADE
 ("VOLATILIDADE — prever o tamanho do sacolejo", None, None, None, None, None, None, None, None, "H"),

 ("Halousková e Lyócsa (2025)", "variância realizada (5 min)", "404 ações do S&P 500",
  "FinBERT + HAR; subconjuntos completos e LASSO adaptativo", "WSJ, FT, Google Trends, Twitter",
  "SUPERAM o HAR em 98,76% dos casos; −12,74% de erro. Maior ganho (14,99%) NOS DIAS EXTREMOS.",
  "SIM", "SIM — mesmo alvo e mesma referência. SUPERAM-NOS. Usam 404 ativos e dados de 5 min.", "V"),

 ("Bodilsen e Lunde (2025)", "volatilidade realizada, vários horizontes", "Ações dos EUA e S&P 500",
  "analítica de notícias comercial + família HAR", "não recuperado",
  "Notícia da EMPRESA não acrescenta; notícia MACRO melhora significativamente, sobretudo em prazo longo.",
  "SIM", "SIM — J. Applied Econometrics. SUPERAM-NOS. Testamos na PETR4 e a hipótese INVERTE-SE.", "V"),

 ("Mino e Williamson (2025)", "volatilidade do índice", "S&P 500 (105 dias)",
  "BERT ajustado + GARCH(1,1)-t — o MESMO do nosso Script 04", "mais de 10.000 manchetes",
  "Coeficiente do sentimento −0,2275 (p=0,0016). O nosso é −0,2924 (p=0,0002).",
  "não testaram", "SIM — magnitude equivalente à nossa. Mas NÃO avaliam fora da amostra; nós avaliamos.", "M"),

 ("Rahimikia e Poon (2021)", "volatilidade realizada", "Ações dos EUA",
  "embedding financeiro próprio (SEM cabeça de sentimento)", "não recuperado",
  "Usam embeddings em vez de classificação de sentimento para prever volatilidade.",
  "parcial", "Indica caminho: usar a COMPREENSÃO do texto, não o parecer positivo/negativo.", "M"),

 ("Horserace cripto (2024)", "volatilidade de criptomoedas", "Criptomoedas",
  "HAR contra LightGBM, XGBoost, LSTM com sentimento", "não recuperado",
  "Sentimento NÃO melhora o HAR linear; melhora com modelos flexíveis (não linearidade).",
  "só com ML", "Sugere que o nosso resultado nulo pode ser de FORMA FUNCIONAL, não de ausência de sinal.", "M"),

 ("Silva (2018)", "retorno e volatilidade", "Índice IBOVESPA (Brasil)",
  "análise de sentimento + GARCH + regressão quantílica", "não declarado",
  "Sentimento linear isolado: R² fora da amostra NEGATIVO. O ganho vem da combinação quantílica.",
  "só quantílico", "SIM — é a nossa base metodológica. Estendemos do índice para ativo individual.", "M"),

 ("A NOSSA — Silva (2026)", "volatilidade da PETR4", "B3 — PETR4 (ação individual)",
  "FinBERT-PT-BR + HAR (Parkinson) + GARCH(1,1)-t", "205.697 manchetes",
  "Coeficiente −0,2924 (p=0,0002), mas NÃO supera o HAR (p=0,64). Melhor recorte: EMPRESA, 22 dias, +1,77% (p=0,0574).",
  "NÃO", "—", "N"),

 # ------------------------------------------------------------------- DIREÇÃO
 ("DIREÇÃO — prever se sobe ou desce", None, None, None, None, None, None, None, None, "H"),

 ("Hashamia e Maldonado (2025)", "DIREÇÃO DA VOLATILIDADE (sobe ou desce)", "Futuros de Brent (petróleo)",
  "VADER, TextBlob, FinBERT, CrudeBERT; embeddings GloVe, FastText, BERT, LLaMA",
  "592.858 manchetes (Reuters)",
  "Contagem de notícias SUPEROU o sentimento. FastText superou as cabeças de sentimento. Código público.",
  "não recuperado", "ALVO QUE NUNCA TESTAMOS — via do meio entre direção do preço e nível da volatilidade. PRIORIDADE 1.", "V"),

 ("Bollen et al. (2011)", "direção do índice", "Índice DJIA (EUA)",
  "OpinionFinder + GPOMS (dicionários, 7 dimensões de humor)", "9.853.498 tuítes",
  "86,7% = 13 ACERTOS EM 15 PREGÕES. REFUTADO por Lachanski e Pav (2017): data snooping.",
  "refutado", "NÃO — 15 dias de teste, índice e não ação, e não replicou. Útil: só 'calma' previu, positivo/negativo não.", "A"),

 ("Schumaker e Chen (2009)", "preço 20 MINUTOS após a notícia", "Ações do S&P 500",
  "saco de palavras, sintagmas nominais, entidades + SVM", "9.211 notícias (5 semanas)",
  "71,18% — o MELHOR entre vários esquemas de particionamento. Retorno simulado de 8,50%.",
  "—", "NÃO — mede REAÇÃO em 20 min (o nosso P0), não o pregão seguinte (P1).", "A"),

 ("Barak et al. (2017)", "retorno e risco", "Bolsa de Teerã",
  "ensembles: bagging, boosting, AdaBoost + meta-classificador", "—",
  "Até 83,6% (retorno) e 88,2% (risco) — o MÁXIMO entre configurações.",
  "—", "NÃO — outro mercado e tarefa. JÁ REPLICAMOS: stacking deu 53,14%, no baseline.", "A"),

 ("Nguyen et al. (2015)", "direção de ações", "Ações dos EUA",
  "TSLDA — sentimento por TÓPICO", "redes sociais",
  "Ganho de 2,1 a 9,8 pontos percentuais sobre a linha de base de preços.",
  "SIM", "SIM — É A COMPARAÇÃO HONESTA. Não a acurácia absoluta, mas o GANHO. O nosso é 4,4 p.p.", "V"),

 ("Li et al. (2020)", "tendência do preço", "Hong Kong",
  "fusão sequencial em aprendizado profundo", "—",
  "Supera as linhas de base ao fundir preços e sentimento.",
  "SIM", "Corrobora a arquitetura de fusão que adotamos.", "M"),

 ("FinBERT-LSTM (2022–2024)", "PREÇO (nível, não direção)", "NASDAQ-100",
  "FinBERT + LSTM", "notícias da Benzinga",
  "Anunciam 'acurácia 0,955' — mas reportam MAE e MAPE, métricas de REGRESSÃO sobre o NÍVEL do preço.",
  "—", "NÃO — 0,955 ≈ 1 − MAPE(0,045). Prever o nível é trivial (autocorrelação ~1). NÃO comparar com 54,5%.", "A"),

 ("FinBERT + SHAP (2025)", "direção do preço", "Ações do S&P 500 (2018–2023)",
  "FinBERT + árvores impulsionadas por gradiente", "manchetes financeiras",
  "Supera linhas de base técnicas e lexicais. Usa SHAP para explicabilidade.",
  "SIM", "Desenho próximo ao nosso (sentimento + preço + volatilidade em árvore). Falta-nos a explicabilidade.", "M"),

 ("A NOSSA — Silva (2026)", "direção da PETR4", "B3 — PETR4 (ação individual)",
  "FinBERT-PT-BR + XGBoost/SVM (3 atributos defasados)", "205.697 manchetes",
  "54,5% de acurácia; GANHO DE 4,4 p.p. sobre o modelo apenas-preços (binomial p=0,012).",
  "SIM", "—", "N"),
]

CORES = {"N": NOSSO, "V": VERDE, "A": VERMELHO, "M": AMARELO}


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Previsão"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
    c = ws.cell(row=1, column=1,
                value="PESQUISAS QUE PREVEEM DIREÇÃO E/OU VOLATILIDADE DE UM ATIVO OU ÍNDICE")
    c.font = Font(bold=True, size=13, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=AZUL)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLS))
    c = ws.cell(row=2, column=1,
                value="Vanderlei Barbosa da Silva · PUCPR/PPGIa · mentoria com o Prof. Dr. Emerson "
                      "Cabrera Paraiso · agosto de 2026")
    c.font = Font(italic=True, size=9, color="404040")
    c.alignment = Alignment(horizontal="center")

    for j, (nome, larg) in enumerate(COLS, start=1):
        c = ws.cell(row=3, column=j, value=nome)
        c.font = Font(bold=True, size=10, color=AZUL)
        c.fill = PatternFill("solid", fgColor=CAB)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = B
        ws.column_dimensions[get_column_letter(j)].width = larg
    ws.row_dimensions[3].height = 32

    n = 0
    for i, item in enumerate(L):
        r = 4 + i
        tipo = item[-1]
        if tipo == "H":                                   # faixa de secao
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(COLS))
            c = ws.cell(row=r, column=1, value=item[0])
            c.font = Font(bold=True, size=11, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=AZUL)
            c.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[r].height = 20
            continue
        n += 1
        ws.cell(row=r, column=1, value=n)
        for j, v in enumerate(item[:-1], start=2):
            ws.cell(row=r, column=j, value=v)
        for j in range(1, len(COLS) + 1):
            c = ws.cell(row=r, column=j)
            c.fill = PatternFill("solid", fgColor=CORES[tipo])
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = B
            c.font = Font(size=9, bold=(tipo == "N" and j == 2))
        ws.row_dimensions[r].height = 76

    ws.freeze_panes = "C4"

    r = 4 + len(L) + 1
    ws.cell(row=r, column=1, value="LEGENDA").font = Font(bold=True, size=10)
    for k, (t, txt) in enumerate([
        ("N", "A nossa pesquisa"),
        ("V", "Resultado superior ao nosso, ou caminho concreto a seguir"),
        ("M", "Comparável, sem superioridade clara"),
        ("A", "ATENÇÃO — número NÃO comparável, refutado, ou já testado sem êxito"),
    ], start=1):
        c = ws.cell(row=r + k, column=1, value="")
        c.fill = PatternFill("solid", fgColor=CORES[t]); c.border = B
        ws.cell(row=r + k, column=2, value=txt).font = Font(size=9)

    r2 = r + 6
    for k, txt in enumerate([
        "LEITURA DE UMA LINHA:",
        "Na VOLATILIDADE, dois trabalhos superam o modelo de referência e nós não — mas eles usam 404 ativos e dados de 5 minutos, contra o nosso 1 ativo e dado diário.",
        "Na DIREÇÃO, ninguém prevê bem: os números altos são de 15 dias (Bollen), de 20 minutos (Schumaker), do nível do preço (FinBERT-LSTM) ou de outro mercado (Barak).",
        "A comparação legítima é o GANHO sobre o modelo só-preços: a literatura reporta 2 a 10 pontos percentuais; o nosso é 4,4. ESTAMOS NA FAIXA.",
    ]):
        c = ws.cell(row=r2 + k, column=1, value=txt)
        c.font = Font(bold=(k == 0), size=10)
        ws.merge_cells(start_row=r2 + k, start_column=1, end_row=r2 + k, end_column=len(COLS))

    wb.save(SAIDA)
    print(f"[OK] {SAIDA}")
    print(f"     {n} pesquisas de previsao (volatilidade e direcao), {len(COLS)} colunas")


if __name__ == "__main__":
    main()
