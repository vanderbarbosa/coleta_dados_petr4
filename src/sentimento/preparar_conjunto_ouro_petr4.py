# -*- coding: utf-8 -*-
# ==============================================================================
#   DISSERTAÇÃO PETR4 — Conjunto-ouro (gold standard) de sentimento
#   Autor: Vanderlei Barbosa da Silva | Orientador: Prof. Dr. Julio Cesar Nievola
#
#   Gera uma AMOSTRA ESTRATIFICADA do corpus para ROTULAGEM HUMANA, com o duplo
#   objetivo de (i) VALIDAR o FinBERT-PT-BR (acurácia e kappa de Cohen contra o
#   rótulo humano) — exigência da banca (Seção 8 da Etapa 3) — e (ii) servir de
#   semente para um futuro AJUSTE FINO do modelo ao corpus desta pesquisa.
#
#   Princípios metodológicos:
#     - Estratificação por classe de sentimento (com piso para a classe minoritária)
#       e por ANO, garantindo cobertura temporal e de classes.
#     - O rótulo do MODELO fica em arquivo SEPARADO (gabarito), para não ancorar
#       (enviesar) o julgamento humano durante a rotulagem.
#     - Peso amostral registrado, permitindo reponderar a acurácia para a
#       distribuição populacional (a amostra superrepresenta a classe minoritária).
#
#   Saídas (em Mestrado_PETR4/conjunto_ouro/):
#     - conjunto_ouro_para_rotular.xlsx   (planilha com menus suspensos)
#     - conjunto_ouro_gabarito_modelo.csv (rótulo do modelo — NÃO abrir antes de rotular)
# ==============================================================================

import os, sys
from pathlib import Path
import numpy as np
import pandas as pd

RAIZ = Path(__file__).resolve().parents[2]
PASTA = RAIZ / "Mestrado_PETR4"
SAIDA = PASTA / "conjunto_ouro"
SAIDA.mkdir(exist_ok=True)

N_TOTAL = int(os.environ.get("CONJUNTO_OURO_N", "300"))   # tamanho da amostra
PISO_CLASSE = 70                                          # mínimo por classe de sentimento
RNG = 42
np.random.seed(RNG)

MAPA_PT = {"Positive": "Positivo", "Negative": "Negativo", "Neutral": "Neutro"}

# ──────────────────────────────────────────────────────────────────────────────
# 1) Carga e limpeza mínima
# ──────────────────────────────────────────────────────────────────────────────
df = pd.read_csv(PASTA / "noticias_com_sentimento.csv")
df = df[df["Titulo"].notna() & (df["Titulo"].str.strip().str.len() >= 15)].copy()
df = df.drop_duplicates(subset="hash_titulo")
df["Ano"] = pd.to_datetime(df["Data"], errors="coerce").dt.year
df = df.dropna(subset=["Ano", "Label_Sentimento"])
df["Ano"] = df["Ano"].astype(int)
N_corpus = len(df)

# ──────────────────────────────────────────────────────────────────────────────
# 2) Alocação por classe de sentimento: piso + proporcional ao restante
# ──────────────────────────────────────────────────────────────────────────────
classes = ["Negative", "Neutral", "Positive"]
freq = df["Label_Sentimento"].value_counts()
prop = (freq / freq.sum())

restante = N_TOTAL - PISO_CLASSE * len(classes)
aloc = {c: PISO_CLASSE + int(round(restante * prop.get(c, 0))) for c in classes}
# ajuste fino para fechar exatamente em N_TOTAL
dif = N_TOTAL - sum(aloc.values())
aloc["Negative"] += dif

# ──────────────────────────────────────────────────────────────────────────────
# 3) Amostragem estratificada por ANO dentro de cada classe
# ──────────────────────────────────────────────────────────────────────────────
def amostra_por_ano(sub, n_alvo):
    """Distribui n_alvo entre os anos proporcionalmente, com pelo menos 1 por ano."""
    anos = sub["Ano"].value_counts()
    pesos = anos / anos.sum()
    aloc_ano = {a: max(1, int(round(n_alvo * pesos[a]))) for a in anos.index}
    partes = []
    for a, na in aloc_ano.items():
        bloco = sub[sub["Ano"] == a]
        partes.append(bloco.sample(min(na, len(bloco)), random_state=RNG))
    out = pd.concat(partes)
    if len(out) > n_alvo:                      # corta excesso aleatoriamente
        out = out.sample(n_alvo, random_state=RNG)
    elif len(out) < n_alvo:                    # completa com o restante da classe
        falta = sub.drop(out.index)
        out = pd.concat([out, falta.sample(min(n_alvo - len(out), len(falta)), random_state=RNG)])
    return out

partes = []
for c in classes:
    sub = df[df["Label_Sentimento"] == c]
    amostra = amostra_por_ano(sub, aloc[c])
    amostra = amostra.copy()
    amostra["peso_amostral"] = len(sub) / len(amostra)   # representa N_classe/n_classe
    partes.append(amostra)

ouro = pd.concat(partes).sample(frac=1, random_state=RNG).reset_index(drop=True)  # embaralha
ouro.insert(0, "ID_OURO", [f"G{ i+1:03d}" for i in range(len(ouro))])

# ──────────────────────────────────────────────────────────────────────────────
# 4) Gabarito do modelo (SEPARADO — não enviesar o rotulador)
# ──────────────────────────────────────────────────────────────────────────────
gabarito = ouro[["ID_OURO", "hash_titulo", "categoria", "Ano", "conjunto",
                 "Label_Sentimento", "Indice_Sentimento", "Score_Confianca",
                 "peso_amostral"]].copy()
gabarito["Sentimento_Modelo_PT"] = gabarito["Label_Sentimento"].map(MAPA_PT)
gabarito.to_csv(SAIDA / "conjunto_ouro_gabarito_modelo.csv", index=False, encoding="utf-8-sig")

# ──────────────────────────────────────────────────────────────────────────────
# 5) Planilha de rotulagem (SEM o rótulo do modelo) com menus suspensos
# ──────────────────────────────────────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

wb = Workbook()

# -- Aba de instruções --
ws0 = wb.active
ws0.title = "Instruções"
instr = [
    ("CONJUNTO-OURO DE SENTIMENTO — PETR4", True),
    ("", False),
    ("Objetivo: rotular manualmente cada notícia para validar o modelo FinBERT-PT-BR", False),
    ("(acurácia e kappa de Cohen) e servir de semente para ajuste fino futuro.", False),
    ("", False),
    ("COMO ROTULAR (aba 'Rotular'):", True),
    ("1. Leia o Título e o Resumo de cada linha.", False),
    ("2. Sentimento_Humano: o tom da notícia para o investidor — Positivo, Negativo ou Neutro.", False),
    ("   (Avalie o SENTIMENTO do texto, não se você concorda com ele.)", False),
    ("3. Relevante_PETR4: a notícia tem relação com a PETR4/Petrobras/petróleo? Sim ou Não.", False),
    ("4. Direcao_Esperada_PETR4: qual o efeito provável no preço da PETR4 — Alta, Baixa ou Indefinida.", False),
    ("   (Atenção: uma notícia pode ser NEGATIVA para o mercado mas de ALTA para a PETR4 —", False),
    ("    ex.: guerra que eleva o preço do petróleo. Use seu julgamento econômico aqui.)", False),
    ("5. Confianca_Rotulador: o quanto você está seguro do rótulo — Alta, Média ou Baixa.", False),
    ("6. Observacao: campo livre (opcional) para casos ambíguos.", False),
    ("", False),
    ("IMPORTANTE: o rótulo do MODELO foi deixado FORA desta planilha de propósito,", True),
    ("para não influenciar (ancorar) o seu julgamento. A comparação será feita depois.", False),
    ("", False),
    (f"Total de notícias a rotular: {N_TOTAL}.  Amostra estratificada por classe e por ano.", False),
]
for i, (txt, bold) in enumerate(instr, start=1):
    c = ws0.cell(row=i, column=1, value=txt)
    c.font = Font(bold=bold, size=13 if (bold and i == 1) else 11)
ws0.column_dimensions["A"].width = 100

# -- Aba de rotulagem --
ws = wb.create_sheet("Rotular")
cols = ["ID_OURO", "Data", "Categoria", "Fonte", "Título", "Resumo", "URL",
        "Sentimento_Humano", "Relevante_PETR4", "Direcao_Esperada_PETR4",
        "Confianca_Rotulador", "Observacao"]
ws.append(cols)
cab_fill = PatternFill("solid", fgColor="0B5394")
for j, _ in enumerate(cols, start=1):
    cell = ws.cell(row=1, column=j)
    cell.font = Font(bold=True, color="FFFFFF"); cell.fill = cab_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for _, r in ouro.iterrows():
    ws.append([r["ID_OURO"], str(r.get("Data", ""))[:10], r.get("categoria", ""),
               r.get("Fonte", ""), r.get("Titulo", ""), r.get("Resumo", ""), r.get("URL", ""),
               "", "", "", "", ""])

larguras = {"A": 9, "B": 12, "C": 20, "D": 14, "E": 60, "F": 70, "G": 30,
            "H": 18, "I": 16, "J": 22, "K": 18, "L": 30}
for col, w in larguras.items():
    ws.column_dimensions[col].width = w
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

# menus suspensos (validação de dados)
def add_dropdown(coluna, opcoes):
    dv = DataValidation(type="list", formula1='"' + ",".join(opcoes) + '"', allow_blank=True)
    dv.error = "Escolha uma opção da lista"; dv.errorTitle = "Valor inválido"
    ws.add_data_validation(dv)
    letra = get_column_letter(coluna)
    dv.add(f"{letra}2:{letra}{len(ouro)+1}")

add_dropdown(8, ["Positivo", "Negativo", "Neutro"])           # Sentimento_Humano
add_dropdown(9, ["Sim", "Não"])                               # Relevante_PETR4
add_dropdown(10, ["Alta", "Baixa", "Indefinida"])             # Direcao_Esperada_PETR4
add_dropdown(11, ["Alta", "Média", "Baixa"])                  # Confianca_Rotulador
ws.freeze_panes = "A2"

wb.save(SAIDA / "conjunto_ouro_para_rotular.xlsx")

# ──────────────────────────────────────────────────────────────────────────────
# 6) Relatório
# ──────────────────────────────────────────────────────────────────────────────
print(f"[OK] Conjunto-ouro gerado em: {SAIDA}")
print(f"     Corpus elegível (dedup, título >=15 chars): {N_corpus:,}".replace(",", "."))
print(f"     Amostra: {len(ouro)} notícias")
print("\n     Alocação por classe de sentimento (modelo):")
for c in classes:
    n_c = (gabarito["Label_Sentimento"] == c).sum()
    print(f"       {MAPA_PT[c]:9s}: {n_c:3d}  (corpus {prop.get(c,0)*100:4.1f}%  |  peso ~{freq[c]/n_c:,.0f})".replace(",", "."))
print("\n     Cobertura por ano:")
print(gabarito["Ano"].value_counts().sort_index().to_string())
print("\n     Cobertura por categoria:")
print(ouro["categoria"].value_counts().to_string())
print("\n     Arquivos:")
print("       - conjunto_ouro_para_rotular.xlsx   (rotular aqui; menus suspensos)")
print("       - conjunto_ouro_gabarito_modelo.csv (NÃO abrir antes de terminar a rotulagem)")
