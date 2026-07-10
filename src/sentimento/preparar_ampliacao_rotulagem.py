# -*- coding: utf-8 -*-
# ==============================================================================
#   DISSERTAÇÃO PETR4 — Planilha de AMPLIAÇÃO do conjunto-ouro (rodada 2)
#   Autor: Vanderlei Barbosa da Silva | Orientador: Prof. Dr. Julio Cesar Nievola
#
#   Seleciona ~400 manchetes NOVAS (fora das 300 já rotuladas) por AMOSTRAGEM POR
#   INCERTEZA (active learning): prioriza os casos de MENOR confiança do FinBERT
#   — onde o modelo mais erra (fronteira neutro↔negativo) — e REFORÇA a classe
#   Positiva (minoritária, 14% do corpus). Gera a planilha de rotulagem (menus
#   suspensos, cega ao modelo) e o gabarito do modelo para posterior avaliação.
# ==============================================================================
import shutil
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

RAIZ = Path(__file__).resolve().parents[2]
OURO = RAIZ / "Mestrado_PETR4" / "conjunto_ouro"
CORPUS = RAIZ / "Mestrado_PETR4" / "noticias_com_sentimento.csv"
QUOTA = {"Neutral": 150, "Negative": 130, "Positive": 120}   # reforça Positivo (corpus ~14%)
SEED = 42

# ── Candidatos: corpus menos os 300 já rotulados ──────────────────────────────
gab0 = pd.read_csv(OURO / "conjunto_ouro_gabarito_modelo.csv")
ja = set(gab0["hash_titulo"])
cols = ["Data_Coleta", "categoria", "Fonte", "dominio", "Titulo", "Resumo", "URL",
        "Idioma", "hash_titulo", "Indice_Sentimento", "Label_Sentimento", "Score_Confianca"]
c = pd.read_csv(CORPUS, usecols=lambda x: x in cols)
c = c[~c["hash_titulo"].isin(ja)].dropna(subset=["Titulo"]).drop_duplicates("hash_titulo")

# ── Amostragem por incerteza (menor confiança) com cotas por classe ───────────
partes = [c[c["Label_Sentimento"] == lab].nsmallest(q, "Score_Confianca") for lab, q in QUOTA.items()]
amost = pd.concat(partes).sample(frac=1, random_state=SEED).reset_index(drop=True)  # embaralha
amost["ID_OURO"] = ["AMP%04d" % (i + 1) for i in range(len(amost))]
print(f"Selecionadas {len(amost)} manchetes | classes(modelo): {amost.Label_Sentimento.value_counts().to_dict()}")
print(f"conf. média das selecionadas: {amost.Score_Confianca.mean():.3f} (mediana do corpus ~0,70)")
print(f"categorias: {amost.categoria.value_counts().to_dict()}")

# gabarito do modelo (para avaliação/treino futuros) — NÃO vai para a aba de rotular
amost[["ID_OURO", "hash_titulo", "categoria", "Label_Sentimento", "Indice_Sentimento",
       "Score_Confianca"]].to_csv(OURO / "gabarito_ampliacao.csv", index=False, encoding="utf-8-sig")

# ── Planilha de rotulagem (cega ao modelo, com menus suspensos) ───────────────
COLS = ["ID_OURO", "Categoria", "Data", "Fonte", "Título", "Resumo", "URL",
        "Sentimento_Humano", "Relevante_PETR4", "Direcao_Esperada_PETR4",
        "Confianca_Rotulador", "Observacao"]
DROPS = {"Sentimento_Humano": "Positivo,Negativo,Neutro", "Relevante_PETR4": "Sim,Não",
         "Direcao_Esperada_PETR4": "Alta,Baixa,Indefinida", "Confianca_Rotulador": "Alta,Média,Baixa"}
LARG = {"ID_OURO": 10, "Categoria": 20, "Data": 12, "Fonte": 14, "Título": 60, "Resumo": 70,
        "URL": 22, "Sentimento_Humano": 18, "Relevante_PETR4": 16, "Direcao_Esperada_PETR4": 20,
        "Confianca_Rotulador": 18, "Observacao": 30}
dados = pd.DataFrame({
    "ID_OURO": amost["ID_OURO"], "Categoria": amost["categoria"], "Data": amost["Data_Coleta"],
    "Fonte": amost["Fonte"], "Título": amost["Titulo"], "Resumo": amost["Resumo"],
    "URL": amost["URL"], "Sentimento_Humano": None, "Relevante_PETR4": None,
    "Direcao_Esperada_PETR4": None, "Confianca_Rotulador": None, "Observacao": None})

wb = Workbook()
ins = wb.active; ins.title = "Instruções"; ins.column_dimensions["A"].width = 110
INSTR = [
    ("AMPLIAÇÃO DO CONJUNTO-OURO (rodada 2) — GUIA DE ROTULAGEM", True),
    ("", False),
    ("Estas 400 manchetes foram escolhidas por AMOSTRAGEM POR INCERTEZA: são os casos em que o", False),
    ("modelo tem MENOS confiança (onde ele mais erra). Rotulá-las é o que mais ajuda a melhorar o", False),
    ("classificador. A rubrica é a MESMA da rodada 1:", False),
    ("", False),
    ("1) Sentimento_Humano — o TOM financeiro (Positivo/Negativo/Neutro), ignorando a Petrobras.", False),
    ("2) Relevante_PETR4 — a notícia plausivelmente afeta a PETR4? (Sim/Não).", False),
    ("3) Direcao_Esperada_PETR4 — efeito no preço: Alta/Baixa/Indefinida (choque de oferta favorece a produtora).", False),
    ("4) Confianca_Rotulador — sua certeza (Alta/Média/Baixa).  5) Observacao — livre.", False),
    ("", False),
    ("Preencha as colunas em amarelo pelo menu suspenso. Salve ao terminar (mesmo arquivo).", False),
]
for i, (t, b) in enumerate(INSTR, 1):
    cel = ins.cell(row=i, column=1, value=t)
    cel.font = Font(bold=b, size=13 if (b and i == 1) else 11, color="0B5394" if b else "222222")
    cel.alignment = Alignment(wrap_text=True, vertical="top")

ws = wb.create_sheet("Rotular")
hf = PatternFill("solid", fgColor="0B5394"); lf = PatternFill("solid", fgColor="FFF2CC")
thin = Border(*(Side(style="thin", color="DDDDDD"),) * 4)
for j, c0 in enumerate(COLS, 1):
    cel = ws.cell(1, j, c0); cel.font = Font(bold=True, color="FFFFFF"); cel.fill = hf
    cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.column_dimensions[get_column_letter(j)].width = LARG.get(c0, 16)
for r_i, (_, row) in enumerate(dados.iterrows(), 2):
    for j, c0 in enumerate(COLS, 1):
        v = row[c0]; cel = ws.cell(r_i, j, None if pd.isna(v) else v)
        cel.alignment = Alignment(wrap_text=(c0 in ("Título", "Resumo")), vertical="top"); cel.border = thin
        if c0 in DROPS: cel.fill = lf
n = len(dados) + 1
for c0, lst in DROPS.items():
    col = get_column_letter(COLS.index(c0) + 1)
    dv = DataValidation(type="list", formula1=f'"{lst}"', allow_blank=True)
    dv.prompt = f"Opções: {lst}"; ws.add_data_validation(dv); dv.add(f"{col}2:{col}{n}")
ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

out = OURO / "rotulagem_ampliacao.xlsx"
wb.save(out)
print(f"✓ Planilha: {out.name} ({len(dados)} manchetes, menus suspensos)")
print(f"✓ Gabarito do modelo: gabarito_ampliacao.csv")
