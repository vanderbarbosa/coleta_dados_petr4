# -*- coding: utf-8 -*-
# ==============================================================================
#   DISSERTAÇÃO PETR4 — Prepara a planilha de ROTULAGEM do conjunto-ouro
#   Autor: Vanderlei Barbosa da Silva | Orientador: Prof. Dr. Julio Cesar Nievola
#
#   Aprimora conjunto_ouro_para_rotular.xlsx com MENUS SUSPENSOS (o rotulador só
#   seleciona), formatação legível e uma aba de instruções/rubrica. Faz backup do
#   original. Rotulagem HUMANA (cega ao modelo) → gabarito para medir acurácia,
#   Kappa e F1 do FinBERT-PT-BR (validação exigida pela banca).
# ==============================================================================
import shutil
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

RAIZ = Path(__file__).resolve().parents[2]
OURO = RAIZ / "Mestrado_PETR4" / "conjunto_ouro"
ARQ = OURO / "conjunto_ouro_para_rotular.xlsx"

orig = pd.read_excel(ARQ, sheet_name="Rotular")
shutil.copyfile(ARQ, OURO / "conjunto_ouro_para_rotular_ORIGINAL.xlsx")

COLS = ["ID_OURO", "Categoria", "Data", "Fonte", "Título", "Resumo", "URL",
        "Sentimento_Humano", "Relevante_PETR4", "Direcao_Esperada_PETR4",
        "Confianca_Rotulador", "Observacao"]
DROPS = {  # coluna -> lista do menu suspenso
    "Sentimento_Humano": "Positivo,Negativo,Neutro",
    "Relevante_PETR4": "Sim,Não",
    "Direcao_Esperada_PETR4": "Alta,Baixa,Indefinida",
    "Confianca_Rotulador": "Alta,Média,Baixa",
}
LARG = {"ID_OURO": 10, "Categoria": 20, "Data": 12, "Fonte": 14, "Título": 60,
        "Resumo": 70, "URL": 22, "Sentimento_Humano": 18, "Relevante_PETR4": 16,
        "Direcao_Esperada_PETR4": 20, "Confianca_Rotulador": 18, "Observacao": 30}

wb = Workbook()

# ── Aba de INSTRUÇÕES / rubrica ───────────────────────────────────────────────
ins = wb.active
ins.title = "Instruções"
ins.column_dimensions["A"].width = 110
INSTR = [
    ("CONJUNTO-OURO DE SENTIMENTO — PETR4 · GUIA DE ROTULAGEM", True),
    ("", False),
    ("Objetivo: criar um gabarito humano para MEDIR a acurácia do modelo de sentimento (FinBERT-PT-BR).", False),
    ("Rotule com base APENAS no texto da manchete/resumo — sem consultar o modelo. Preencha as colunas de menu.", False),
    ("", False),
    ("1) Sentimento_Humano — o TOM FINANCEIRO da notícia (não a sua opinião):", True),
    ("   • Positivo: sugere melhora/valorização (lucro, alta, acordo, dividendos, descoberta, expansão).", False),
    ("   • Negativo: sugere piora/desvalorização (prejuízo, queda, greve, sanção, acidente, corrupção).", False),
    ("   • Neutro: factual, sem valência clara (agenda, nomeação técnica, dado misto/ambíguo).", False),
    ("   Dica: pergunte 'isto tende a AGRADAR ou DESAGRADAR o mercado?'. Se não der para dizer, é Neutro.", False),
    ("", False),
    ("2) Relevante_PETR4 — a notícia PLAUSIVELMENTE afeta o preço da PETR4? (Sim/Não)", True),
    ("   • Sim: fala da Petrobras, do petróleo/Brent, de choques de oferta, geopolítica do petróleo, câmbio/juros relevantes.", False),
    ("   • Não: tangencial (ex.: empresa homônima, assunto sem ligação com o ativo).", False),
    ("", False),
    ("3) Direcao_Esperada_PETR4 — pela leitura econômica, o efeito no PREÇO da PETR4 tende a ser:", True),
    ("   • Alta / Baixa / Indefinida. Lembre: choque de oferta (guerra, sanção) tende a ELEVAR o petróleo e FAVORECER a produtora,", False),
    ("     mesmo que o tom seja negativo para o mercado em geral (Kilian, 2009; Hamilton, 1983).", False),
    ("", False),
    ("4) Confianca_Rotulador — o seu grau de certeza (Alta/Média/Baixa). 5) Observacao — livre (opcional).", True),
    ("", False),
    ("Ao terminar, salve o arquivo. O script src/sentimento/avaliar_conjunto_ouro_petr4.py calcula", False),
    ("acurácia, Kappa de Cohen, F1 por classe e matriz de confusão (modelo × humano).", False),
]
for i, (txt, bold) in enumerate(INSTR, start=1):
    cel = ins.cell(row=i, column=1, value=txt)
    cel.font = Font(bold=bold, size=13 if (bold and i == 1) else 11,
                    color="0B5394" if bold else "222222")
    cel.alignment = Alignment(wrap_text=True, vertical="top")

# ── Aba ROTULAR ───────────────────────────────────────────────────────────────
ws = wb.create_sheet("Rotular")
head_fill = PatternFill("solid", fgColor="0B5394")
lab_fill = PatternFill("solid", fgColor="FFF2CC")   # colunas a preencher (amarelo)
thin = Border(*(Side(style="thin", color="DDDDDD"),) * 4)
for j, c in enumerate(COLS, start=1):
    cel = ws.cell(row=1, column=j, value=c)
    cel.font = Font(bold=True, color="FFFFFF"); cel.fill = head_fill
    cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.column_dimensions[get_column_letter(j)].width = LARG.get(c, 16)

for r_i, (_, row) in enumerate(orig.iterrows(), start=2):
    for j, c in enumerate(COLS, start=1):
        val = row[c] if c in orig.columns else None
        cel = ws.cell(row=r_i, column=j, value=(None if pd.isna(val) else val))
        cel.alignment = Alignment(wrap_text=(c in ("Título", "Resumo")), vertical="top")
        cel.border = thin
        if c in DROPS:
            cel.fill = lab_fill

# menus suspensos
n = len(orig) + 1
for c, lista in DROPS.items():
    col = get_column_letter(COLS.index(c) + 1)
    dv = DataValidation(type="list", formula1=f'"{lista}"', allow_blank=True, showDropDown=False)
    dv.error = "Selecione uma opção da lista."; dv.prompt = f"Opções: {lista}"
    ws.add_data_validation(dv)
    dv.add(f"{col}2:{col}{n}")

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

wb.save(ARQ)
print(f"✓ Planilha de rotulagem pronta: {ARQ}")
print(f"  {len(orig)} manchetes | menus suspensos: {', '.join(DROPS)}")
print(f"  Backup do original: conjunto_ouro_para_rotular_ORIGINAL.xlsx")
